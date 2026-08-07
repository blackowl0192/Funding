from __future__ import annotations

import csv
import io
import logging
import re
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import PurePath

from openpyxl import load_workbook

from funding_terminal.config import Settings
from funding_terminal.domain.enums import ImportRunStatus, MappingStatus
from funding_terminal.domain.errors import ImportValidationError
from funding_terminal.domain.models import AssetInput, ImportResult, UniverseEntry
from funding_terminal.services.universe_service import UniverseMatcher

logger = logging.getLogger(__name__)

MAX_IMPORT_BYTES = 2 * 1024 * 1024
BASE_ASSET_RE = re.compile(r"^[A-Z0-9]{2,20}$")
PAIR_QUOTE_SUFFIXES = ("USDC", "FDUSD", "BUSD", "BTC", "ETH", "BNB", "EUR", "TRY")


@dataclass(frozen=True, slots=True)
class ParsedUniverse:
    file_type: str
    total_rows: int
    assets: tuple[AssetInput, ...]
    invalid_entries: tuple[UniverseEntry, ...]
    duplicate_count: int

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_entries)


class ImportFileParser:
    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    def parse(self, filename: str, content: bytes) -> ParsedUniverse:
        safe_name = safe_filename(filename)
        if not safe_name:
            raise ImportValidationError("Import filename is required.")
        if not content:
            raise ImportValidationError("Import file is empty.")
        if len(content) > MAX_IMPORT_BYTES:
            raise ImportValidationError("Import file is too large. Maximum size is 2 MB.")

        suffix = PurePath(safe_name).suffix.lower()
        if suffix == ".csv":
            rows = self._read_csv(content)
            return self._parse_rows("CSV", rows)
        if suffix == ".xlsx":
            rows = self._read_xlsx(content)
            return self._parse_rows("XLSX", rows)
        raise ImportValidationError("Only CSV and XLSX files are supported.")

    def _read_csv(self, content: bytes) -> list[dict[str, str]]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ImportValidationError("CSV must be UTF-8 encoded.") from exc

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ImportValidationError("CSV must include a header row with Symbol column.")

        symbol_key = _header_key(reader.fieldnames, "Symbol")
        if symbol_key is None:
            raise ImportValidationError("Import file must include a Symbol column.")
        enabled_key = _header_key(reader.fieldnames, "Enabled")

        rows: list[dict[str, str]] = []
        for row in reader:
            rows.append(
                {
                    "Symbol": row.get(symbol_key, "") or "",
                    "Enabled": row.get(enabled_key, "") if enabled_key is not None else "",
                }
            )
        return rows

    def _read_xlsx(self, content: bytes) -> list[dict[str, str]]:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ImportValidationError("XLSX file could not be read.") from exc

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration as exc:
            message = "XLSX must include a header row with Symbol column."
            raise ImportValidationError(message) from exc

        header_values = ["" if value is None else str(value) for value in headers]
        symbol_index = _header_index(header_values, "Symbol")
        if symbol_index is None:
            raise ImportValidationError("Import file must include a Symbol column.")
        enabled_index = _header_index(header_values, "Enabled")

        rows: list[dict[str, str]] = []
        for row in rows_iter:
            symbol = _cell(row, symbol_index)
            enabled = _cell(row, enabled_index) if enabled_index is not None else ""
            rows.append({"Symbol": symbol, "Enabled": enabled})
        return rows

    def _parse_rows(self, file_type: str, rows: list[dict[str, str]]) -> ParsedUniverse:
        if len(rows) > self._settings.import_max_rows:
            raise ImportValidationError(
                f"Import has {len(rows)} rows, maximum is {self._settings.import_max_rows}."
            )

        seen: set[str] = set()
        assets: list[AssetInput] = []
        invalid_entries: list[UniverseEntry] = []
        duplicate_count = 0

        for row_number, row in enumerate(rows, start=2):
            raw_symbol = row.get("Symbol", "")
            enabled = parse_enabled(row.get("Enabled", ""))
            normalized = normalize_symbol(raw_symbol)
            if normalized is None:
                continue
            base_asset, error = normalized
            if error is not None:
                invalid_entries.append(_invalid_entry(base_asset, error, enabled))
                continue
            if base_asset in seen:
                duplicate_count += 1
                continue
            seen.add(base_asset)
            assets.append(
                AssetInput(
                    raw_symbol=raw_symbol,
                    base_asset=base_asset,
                    enabled=enabled,
                    row_number=row_number,
                )
            )

        return ParsedUniverse(
            file_type=file_type,
            total_rows=len(rows),
            assets=tuple(assets),
            invalid_entries=tuple(invalid_entries),
            duplicate_count=duplicate_count,
        )


class ImportService:
    def __init__(
        self,
        settings: Settings,
        symbol_repository,
        import_repository,
        binance_client,
    ) -> None:
        self._parser = ImportFileParser(settings)
        self._symbol_repository = symbol_repository
        self._import_repository = import_repository
        self._binance_client = binance_client
        self._matcher = UniverseMatcher()

    async def import_universe(self, filename: str, content: bytes) -> ImportResult:
        safe_name = safe_filename(filename)
        logger.info("import started: %s", safe_name)
        try:
            parsed = self._parser.parse(safe_name, content)
            metadata = await self._binance_client.load_exchange_metadata()
            matched_entries = self._matcher.match_assets(parsed.assets, metadata)
            entries = (*matched_entries, *parsed.invalid_entries)
            persistable_entries = tuple(
                entry
                for entry in entries
                if entry.mapping_status != MappingStatus.INVALID_INPUT or entry.base_asset.isalnum()
            )
            if persistable_entries:
                await self._symbol_repository.upsert_many(persistable_entries)

            matched_count = sum(1 for entry in entries if entry.strategy_eligible)
            rejected_count = len(entries) - matched_count
            result = ImportResult(
                filename=safe_name,
                file_type=parsed.file_type,
                total_rows=parsed.total_rows,
                unique_assets=len(parsed.assets),
                matched_count=matched_count,
                rejected_count=rejected_count,
                duplicate_count=parsed.duplicate_count,
                invalid_count=parsed.invalid_count,
                entries=entries,
            )
            await self._import_repository.record_result(result, ImportRunStatus.COMPLETED)
            logger.info("import completed: %s", safe_name)
            return result
        except Exception as exc:
            await self._import_repository.record_failure(safe_name, str(exc))
            logger.exception("import failed: %s", safe_name)
            raise


def safe_filename(filename: str | None) -> str:
    if not filename:
        return ""
    return PurePath(filename).name


def parse_enabled(value: str | None) -> bool:
    if value is None or value.strip() == "":
        return True
    normalized = value.strip().upper()
    if normalized in {"1", "TRUE", "YES", "Y", "ENABLED", "ON"}:
        return True
    return normalized not in {"0", "FALSE", "NO", "N", "DISABLED", "OFF"}


def normalize_symbol(raw_symbol: str | None) -> tuple[str, str | None] | None:
    cleaned = (raw_symbol or "").replace("\ufeff", "").strip().strip('"').strip("'").upper()
    cleaned = cleaned.replace(" ", "")
    if not cleaned:
        return None

    if cleaned.endswith("USDT") and len(cleaned) > 4:
        base_asset = cleaned[:-4]
    elif any(
        cleaned.endswith(suffix) and len(cleaned) > len(suffix)
        for suffix in PAIR_QUOTE_SUFFIXES
    ):
        return cleaned, "Only BASE or BASEUSDT inputs are accepted in Stage 1."
    else:
        base_asset = cleaned

    if not BASE_ASSET_RE.fullmatch(base_asset):
        return base_asset or cleaned, "Symbol must contain 2-20 uppercase letters or digits."
    return base_asset, None


def _invalid_entry(base_asset: str, reason: str, enabled: bool) -> UniverseEntry:
    return UniverseEntry(
        base_asset=base_asset,
        spot_symbol=None,
        futures_symbol=None,
        spot_status=None,
        futures_status=None,
        mapping_status=MappingStatus.INVALID_INPUT,
        mapping_reason=reason,
        strategy_eligible=False,
        enabled=enabled,
    )


def _header_key(headers: Sequence[str], expected: str) -> str | None:
    expected_lower = expected.lower()
    for header in headers:
        if header.replace("\ufeff", "").strip().lower() == expected_lower:
            return header
    return None


def _header_index(headers: Sequence[str], expected: str) -> int | None:
    expected_lower = expected.lower()
    for index, header in enumerate(headers):
        if header.replace("\ufeff", "").strip().lower() == expected_lower:
            return index
    return None


def _cell(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value)
